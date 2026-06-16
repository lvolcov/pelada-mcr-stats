"""Excel parser for the Pelada match-tracking workbook.

Reads the raw ``Player Match Stats`` sheet (the single source of truth) and the
``Jogadores`` registry, returning clean Python structures. Results are cached and
invalidated automatically when the file's modification time changes, so dropping a
new workbook into the data folder refreshes everything on the next request.

Created: 2026-06-16
"""

from __future__ import annotations

import json
import threading
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path

import openpyxl

RAW_SHEET = "Player Match Stats"
PLAYERS_SHEET = "Jogadores"


@dataclass
class MatchRow:
    """One player's record in a single session."""

    date: date
    score: str
    player: str
    goals: int
    assists: int
    win: int
    loss: int
    draw: int
    mixed: int  # "Time misto" — teams were reshuffled (3-team day)


@dataclass
class Dataset:
    """Everything parsed from one workbook snapshot."""

    rows: list[MatchRow] = field(default_factory=list)
    registered_players: list[str] = field(default_factory=list)
    # player name (lowercase) -> ISO date they became a mensalista (fixed weekly spot)
    mensalistas: dict[str, str] = field(default_factory=dict)
    source_mtime: float = 0.0
    parsed_at: str = ""


def _to_int(value) -> int:
    """Coerce a cell to int, treating blanks/None as 0."""
    if value is None or value == "":
        return 0
    try:
        return int(value)
    except (TypeError, ValueError):
        return 0


def _to_date(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def _norm_score(value) -> str:
    """Normalise score text, e.g. '5 X 3 ' -> '5 x 3'."""
    if value is None:
        return ""
    return str(value).strip().replace(" X ", " x ").replace("X", "x") \
        if any(c.isdigit() for c in str(value)) else str(value).strip()


class WorkbookStore:
    """Thread-safe, mtime-cached loader for the workbook."""

    def __init__(self, path: str | Path, mensalistas_path: str | Path | None = None) -> None:
        self._path = Path(path)
        # Mensalistas registry lives next to the workbook by default.
        self._mensalistas_path = (
            Path(mensalistas_path)
            if mensalistas_path
            else self._path.parent / "mensalistas.json"
        )
        self._lock = threading.Lock()
        self._cache: Dataset | None = None

    @property
    def path(self) -> Path:
        return self._path

    def exists(self) -> bool:
        return self._path.exists()

    def get(self) -> Dataset:
        """Return the parsed dataset, re-parsing if the file changed."""
        if not self._path.exists():
            raise FileNotFoundError(f"Workbook not found at {self._path}")

        # Cache key combines both files' mtimes so editing either refreshes.
        mtime = self._path.stat().st_mtime
        if self._mensalistas_path.exists():
            mtime += self._mensalistas_path.stat().st_mtime
        with self._lock:
            if self._cache is None or self._cache.source_mtime != mtime:
                self._cache = self._parse(mtime)
            return self._cache

    def _load_mensalistas(self) -> dict[str, str]:
        if not self._mensalistas_path.exists():
            return {}
        try:
            raw = json.loads(self._mensalistas_path.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            return {}
        data = raw.get("mensalistas", raw) if isinstance(raw, dict) else {}
        # Normalise keys to lowercase; drop any metadata keys like "_note".
        return {
            str(k).strip().lower(): str(v)
            for k, v in data.items()
            if not str(k).startswith("_")
        }

    def _parse(self, mtime: float) -> Dataset:
        wb = openpyxl.load_workbook(self._path, data_only=True, read_only=True)

        rows: list[MatchRow] = []
        ws = wb[RAW_SHEET]
        for raw in ws.iter_rows(min_row=2, values_only=True):
            d = _to_date(raw[0])
            player = raw[2]
            if d is None or not player:
                continue
            rows.append(
                MatchRow(
                    date=d,
                    score=_norm_score(raw[1]),
                    player=str(player).strip().lower(),
                    goals=_to_int(raw[3]),
                    assists=_to_int(raw[4]),
                    win=_to_int(raw[5]),
                    loss=_to_int(raw[6]),
                    draw=_to_int(raw[7]),
                    mixed=_to_int(raw[8]),
                )
            )

        registered: list[str] = []
        if PLAYERS_SHEET in wb.sheetnames:
            pws = wb[PLAYERS_SHEET]
            for raw in pws.iter_rows(min_row=2, values_only=True):
                name = raw[0]
                if name and str(name).strip():
                    registered.append(str(name).strip().lower())

        wb.close()
        return Dataset(
            rows=rows,
            registered_players=registered,
            mensalistas=self._load_mensalistas(),
            source_mtime=mtime,
            parsed_at=datetime.now().isoformat(timespec="seconds"),
        )
